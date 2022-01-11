import argparse
import netaddr
import openpyxl

def test():
    path_xls = "test1.xlsx"
    book = openpyxl.load_workbook(path_xls)
    sheet = book.active
    start = input("変換したい始めのcellNo:")
    end = input("変換したい終わりのcellNo:")
    output_C = input("出力したい列:")
    cells = sheet[start:end]
    msg = "IPレンジが正しくないため、CIDR表記できません。"

    #create output file
    output_wb = "output.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    for colums in list(cells):
        for cell in colums:
            if "-" not in cell.value:
#                print(netaddr.iprange_to_cidrs(cell.value,cell.value))
                ip = netaddr.iprange_to_cidrs(cell.value,cell.value)
                ws[cell.coordinate] = cell.value
                ip = str(ip)
                ws[output_C + cell.coordinate[1:]] = ip
            else:
                st,en = cell.value.split("-")
                ip = netaddr.iprange_to_cidrs(st,en)
                ws[cell.coordinate] = cell.value
                if len(ip) == 1:
                    ip = str(ip)
                    ws[output_C + cell.coordinate[1:]] = ip
                else:
                    ws[output_C + cell.coordinate[1:]] = msg

    wb.save(output_wb)

def main():
    parser = argparse.ArgumentParser(description="IPアドレスをCIDR表記にするプログラム")

    parser.add_argument("filename",help="読み込むブック名")
    parser.add_argument('startcellno', help='読込むセル名(例 A1): ')
    parser.add_argument('endcellno', help='読込むセル名(例 A10): ')

    args = parser.parse_args()
    book = openpyxl.load_workbook(args.filename)
    sheet = book.active

    cells = sheet[args.startcellno:args.endcellno]
    for colums in list(cells):
        for cell in colums:
    #        print(cell.value)
            if "-" not in cell.value:
                print(netaddr.iprange_to_cidrs(cell.value,cell.value))
    #book.save("hogehoge.xlsx")

if __name__ == "__main__":
#    main()
    test()